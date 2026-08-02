import ctypes
import time
import sys
import re
import threading
import numpy as np
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QLineEdit, QGroupBox, QFileDialog,
    QStatusBar, QTabWidget, QProgressBar, QSizePolicy,
    QCheckBox, QScrollArea, QFrame, QComboBox,
    QDoubleSpinBox, QSpinBox, QListWidget, QAbstractItemView,
    QMessageBox, QTableWidget, QTableWidgetItem, QHeaderView,
)
from PyQt6.QtCore import QThread, pyqtSignal, Qt
import matplotlib
matplotlib.use("QtAgg")
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.patches import Rectangle as MplRect
from scipy.signal import savgol_filter, find_peaks

import avaspec

COLORS = ["#4da6ff", "#f0a500", "#50fa7b", "#ff79c6", "#bd93f9", "#8be9fd", "#ffb86c"]


def make_collapsible(groupbox, start_expanded=True):
    """
    Convierte un QGroupBox en colapsable: aparece una casilla en el titulo
    que, al desmarcarla, oculta todo el contenido del grupo (dejando solo
    la cabecera) para ahorrar espacio vertical; al volver a marcarla, se
    muestra de nuevo. Debe llamarse DESPUES de rellenar por completo el
    layout del groupbox con sus widgets.
    """
    groupbox.setCheckable(True)
    groupbox.setChecked(start_expanded)

    def _toggle(checked, gb=groupbox):
        layout = gb.layout()
        if layout is None:
            return
        for i in range(layout.count()):
            item = layout.itemAt(i)
            w = item.widget()
            if w is not None:
                w.setVisible(checked)
                continue
            sub = item.layout()
            if sub is not None:
                for j in range(sub.count()):
                    sw = sub.itemAt(j).widget()
                    if sw is not None:
                        sw.setVisible(checked)

    groupbox.toggled.connect(_toggle)
    _toggle(start_expanded)
    return groupbox

ELEMENTS = [
    "Ag","Al","Ar","As","Au","B","Ba","Be","Bi","Br","C","Ca","Cd","Ce","Cl",
    "Co","Cr","Cs","Cu","Dy","Er","Eu","F","Fe","Ga","Gd","Ge","H","He","Hf",
    "Hg","Ho","I","In","Ir","K","Kr","La","Li","Lu","Mg","Mn","Mo","N","Na",
    "Nb","Nd","Ne","Ni","O","Os","P","Pb","Pd","Pr","Pt","Rb","Re","Rh","Ru",
    "S","Sb","Sc","Se","Si","Sm","Sn","Sr","Ta","Tb","Tc","Te","Th","Ti","Tl",
    "Tm","U","V","W","Xe","Y","Yb","Zn","Zr"
]


# ===============================================
# CONVERSION DE UNIDADES DE TIEMPO
# ===============================================
UNIT_TO_MS = {
    "ms": 1.0,
    "us": 1e-3,
    "ns": 1e-6,
}


def to_ms(value_text, unit):
    """Convierte un valor numerico (string) en la unidad dada (ms/us/ns) a milisegundos."""
    try:
        v = float(value_text)
    except ValueError:
        v = 0.0
    return v * UNIT_TO_MS.get(unit, 1.0)


# ===============================================
# ESPECTROMETRO
# ===============================================
class Spectrometer:
    def __init__(self, log, identity):
        self.log = log
        self.identity = identity
        self.handle = None
        self.connected = False
        self.n_pixels = 0
        self.wl = None
        self.last_t = None

    def connect(self):
        self.handle = avaspec.AVS_Activate(self.identity)
        if self.handle <= 0:
            self.log("Error al activar canal"); return False
        ret_hr = avaspec.AVS_UseHighResAdc(self.handle, True)
        self.log(f"AVS_UseHighResAdc ret={ret_hr}")
        pixels = ctypes.c_uint16()
        avaspec.AVS_GetNumPixels(self.handle, ctypes.byref(pixels))
        self.n_pixels = pixels.value
        wl_buf = (ctypes.c_double * self.n_pixels)()
        avaspec.AVS_GetLambda(self.handle, wl_buf)
        self.wl = np.array(wl_buf[:self.n_pixels])
        if np.max(self.wl) < 100:
            self.wl = np.arange(self.n_pixels)
        self.connected = True
        return True

    def disconnect(self):
        try:
            avaspec.AVS_Deactivate(self.handle)
        except Exception:
            pass
        self.connected = False

    def prepare_hw(self, t_ms, edge=0, integration_delay_us=0):
        meas = avaspec.MeasConfigType()
        meas.m_StartPixel           = 0
        meas.m_StopPixel            = self.n_pixels - 1
        meas.m_IntegrationTime      = float(t_ms)
        meas.m_IntegrationDelay     = int(integration_delay_us)
        meas.m_NrAverages           = 1
        meas.m_SaturationDetection  = 1   # rango dinamico extendido (igual que AvaSoft)
        meas.m_Trigger_m_Mode       = edge   # 0=SW interno, 1=HW externo
        meas.m_Trigger_m_Source     = 0
        meas.m_Trigger_m_SourceType = 0
        ret = avaspec.AVS_PrepareMeasure(self.handle, meas)
        self.last_t = None
        self.log(f"AVS_PrepareMeasure ret={ret}")
        self.log(f"[CONFIG] StartPixel={meas.m_StartPixel}  StopPixel={meas.m_StopPixel}")
        self.log(f"[CONFIG] IntegrationTime={meas.m_IntegrationTime} ms  IntegrationDelay={meas.m_IntegrationDelay} us")
        self.log(f"[CONFIG] NrAverages={meas.m_NrAverages}")
        self.log(f"[CONFIG] Trigger Mode={meas.m_Trigger_m_Mode}  Source={meas.m_Trigger_m_Source}  SourceType={meas.m_Trigger_m_SourceType}")
        self.log(f"[CONFIG] DynDark Enable={meas.m_CorDynDark_m_Enable}  ForgetPct={meas.m_CorDynDark_m_ForgetPercentage}")
        self.log(f"[CONFIG] Smoothing Pix={meas.m_Smoothing_m_SmoothPix}  Model={meas.m_Smoothing_m_SmoothModel}")
        self.log(f"[CONFIG] SaturationDetection={meas.m_SaturationDetection}")
        self.log(f"[CONFIG] StrobeControl={meas.m_Control_m_StrobeControl}")
        self.log(f"[CONFIG] LaserDelay={meas.m_Control_m_LaserDelay}  LaserWidth={meas.m_Control_m_LaserWidth}")
        self.log(f"[CONFIG] LaserWaveLength={meas.m_Control_m_LaserWaveLength}  StoreToRam={meas.m_Control_m_StoreToRam}")

    def arm(self):
        ret = avaspec.AVS_Measure(self.handle, None, 1)
        self.log(f"AVS_Measure ret={ret}")

    def read(self, timeout=60):
        raw   = (ctypes.c_double * self.n_pixels)()
        ts    = ctypes.c_uint32()
        start = time.time()
        while not avaspec.AVS_PollScan(self.handle):
            if time.time() - start > timeout:
                return None
            time.sleep(0.002)
        avaspec.AVS_GetScopeData(self.handle, ctypes.byref(ts), raw)
        return np.array(raw)


# ===============================================
# PROCESADO
# ===============================================
class Processor:
    @staticmethod
    def process(s, normalize=True, do_filter=True, window=11, poly=3,
                clip_upper_enabled=False, clip_upper_value=None,
                clip_lower_enabled=False, clip_lower_value=None):
        s = s - np.min(s)
        if normalize and np.max(s) > 0:
            s = s / np.max(s)
        if do_filter:
            n = len(s)
            w = int(window)
            if w % 2 == 0:
                w += 1
            if w < 3:
                w = 3
            p = int(poly)
            if p < 0:
                p = 0
            if w <= p:
                w = p + 2 if (p + 2) % 2 == 1 else p + 3
            if w >= n:
                w = n - 1 if (n - 1) % 2 == 1 else n - 2
            if w > p and w >= 3 and w < n:
                s = savgol_filter(s, w, p)
        # filtro de umbral: recorta (satura) picos por encima/por debajo
        # de un valor dado, sin alterar el resto del espectro
        if clip_upper_enabled and clip_upper_value is not None:
            s = np.minimum(s, float(clip_upper_value))
        if clip_lower_enabled and clip_lower_value is not None:
            s = np.maximum(s, float(clip_lower_value))
        return s

    @staticmethod
    def merge_channels(wavelengths, spectra, n_points=None, overlap_frac=0.08):
        """
        Combina varios canales de espectrometro en un unico eje de longitud
        de onda, fusionando suavemente las zonas de solape (en vez de
        simplemente concatenar y ordenar, lo que producia escalones/duplicados
        bruscos en las regiones donde dos canales cubren la misma zona).

        Cada canal recibe un peso tipo "rampa" (0->1->0) a lo largo de su
        propio rango: cae a 0 en sus extremos y vale 1 en su zona central,
        de forma que en una region de solape entre canal A (cayendo) y canal
        B (subiendo) el resultado es una mezcla ponderada y continua entre
        ambos en vez de un salto.

        n_points: numero de puntos del eje combinado. Si se deja en None
        (por defecto) se calcula automaticamente como la suma de los
        pixeles de todos los canales, para no perder resolucion cuando hay
        varios espectrometros encadenados (p.ej. 3 canales de 2048 pixeles
        deberian dar ~6144 puntos combinados, no 2048).
        """
        if len(wavelengths) == 1:
            return wavelengths[0], spectra[0]

        if n_points is None:
            n_points = int(sum(len(wl) for wl in wavelengths))

        lo = min(float(np.min(wl)) for wl in wavelengths)
        hi = max(float(np.max(wl)) for wl in wavelengths)
        wl_c = np.linspace(lo, hi, n_points)

        acc  = np.zeros(n_points)
        wsum = np.zeros(n_points)

        for wl, s in zip(wavelengths, spectra):
            wl0, wl1 = float(wl[0]), float(wl[-1])
            span = max(wl1 - wl0, 1e-9)
            edge = span * overlap_frac

            s_i = np.interp(wl_c, wl, s, left=np.nan, right=np.nan)
            in_range = ~np.isnan(s_i)

            left_ramp  = np.clip((wl_c - wl0) / edge, 0.0, 1.0)
            right_ramp = np.clip((wl1 - wl_c) / edge, 0.0, 1.0)
            weight = np.minimum(left_ramp, right_ramp)
            weight = np.where(in_range, weight, 0.0)

            s_i = np.nan_to_num(s_i)
            acc  += s_i * weight
            wsum += weight

        # puntos sin ningun canal con peso > 0 (huecos) -> reparto uniforme
        zero_mask = wsum <= 1e-9
        if np.any(zero_mask):
            for wl, s in zip(wavelengths, spectra):
                s_i = np.interp(wl_c, wl, s, left=np.nan, right=np.nan)
                in_range = (~np.isnan(s_i)) & zero_mask
                acc[in_range]  += np.nan_to_num(s_i)[in_range]
                wsum[in_range] += 1.0

        wsum[wsum <= 1e-9] = 1.0
        s_out = acc / wsum
        return wl_c, s_out

    @staticmethod
    def apply_view_filters(s, do_sg=False, window=11, poly=3,
                            clip_upper_enabled=False, clip_upper_value=None,
                            clip_lower_enabled=False, clip_lower_value=None):
        """
        Filtrado 'a posteriori' para la vista de Comparative: NO reajusta el
        offset ni normaliza (los datos guardados/exportados no se tocan),
        solo aplica opcionalmente Savitzky-Golay y/o recorte de umbrales
        superior/inferior sobre una copia, para visualizar el efecto.
        """
        s = np.asarray(s, dtype=float).copy()
        if do_sg:
            n = len(s)
            w = int(window)
            if w % 2 == 0:
                w += 1
            if w < 3:
                w = 3
            p = int(poly)
            if p < 0:
                p = 0
            if w <= p:
                w = p + 2 if (p + 2) % 2 == 1 else p + 3
            if w >= n:
                w = n - 1 if (n - 1) % 2 == 1 else n - 2
            if w > p and w >= 3 and w < n:
                s = savgol_filter(s, w, p)
        if clip_upper_enabled and clip_upper_value is not None:
            s = np.minimum(s, float(clip_upper_value))
        if clip_lower_enabled and clip_lower_value is not None:
            s = np.maximum(s, float(clip_lower_value))
        return s


# ===============================================
# HILO DE MEDIDA
# ===============================================
class TriggerThread(QThread):
    shot_received = pyqtSignal(int, object, object)
    finished_all  = pyqtSignal()

    def __init__(self, specs, t_ms, n_shots=0, edge=0,
                 normalize=False, accum_mode=0, integration_delay_us=0,
                 do_filter=True, filter_window=11, filter_poly=3,
                 clip_upper_enabled=False, clip_upper_value=None,
                 clip_lower_enabled=False, clip_lower_value=None):
        super().__init__()
        self.specs               = specs
        self.t_ms                = t_ms
        self.n_shots             = n_shots
        self.edge                = edge
        self.normalize           = normalize
        self.accum_mode          = accum_mode
        self.integration_delay_us = int(integration_delay_us)
        self.do_filter           = do_filter
        self.filter_window       = filter_window
        self.filter_poly         = filter_poly
        self.clip_upper_enabled  = clip_upper_enabled
        self.clip_upper_value    = clip_upper_value
        self.clip_lower_enabled  = clip_lower_enabled
        self.clip_lower_value    = clip_lower_value
        self._running            = True
        self._shot_idx           = 0
        self._accum              = None

    def run(self):
        for spec in self.specs:
            spec.prepare_hw(self.t_ms, self.edge, self.integration_delay_us)

        while self._running:
            if self.n_shots > 0 and self._shot_idx >= self.n_shots:
                break

            def _arm_all(specs):
                threads = [threading.Thread(target=spec.arm) for spec in specs]
                for th in threads: th.start()
                for th in threads: th.join()

            _arm_all(self.specs)

            spectra, wavelengths = [], []
            for spec in self.specs:
                s = spec.read(timeout=60)
                if s is None:
                    continue
                raw_max = float(np.max(s))
                spec.log(f"[DEBUG] Max crudo (sin procesar): {raw_max:.1f}")
                s = Processor.process(
                    s, normalize=self.normalize, do_filter=self.do_filter,
                    window=self.filter_window, poly=self.filter_poly,
                    clip_upper_enabled=self.clip_upper_enabled,
                    clip_upper_value=self.clip_upper_value,
                    clip_lower_enabled=self.clip_lower_enabled,
                    clip_lower_value=self.clip_lower_value,
                )
                spectra.append(s)
                wavelengths.append(spec.wl)

            if not spectra:
                continue

            wl_c, s_out = Processor.merge_channels(wavelengths, spectra)

            if self.accum_mode == 1:
                self._accum = s_out.copy() if self._accum is None \
                              else self._accum + s_out
                emit_s = self._accum
            elif self.accum_mode == 2:
                n = self._shot_idx + 1
                self._accum = s_out.copy() if self._accum is None \
                              else self._accum * (n - 1) / n + s_out / n
                emit_s = self._accum
            else:
                emit_s = s_out

            self._shot_idx += 1
            self.shot_received.emit(self._shot_idx, wl_c, emit_s)

        self.finished_all.emit()

    def stop(self):
        self._running = False

    def update_filter_params(self, do_filter, filter_window, filter_poly,
                              clip_upper_enabled, clip_upper_value,
                              clip_lower_enabled, clip_lower_value):
        """
        Actualiza en caliente los parametros de filtrado (Savitzky-Golay y
        umbrales de recorte) mientras la adquisicion esta en marcha. El
        bucle de 'run()' relee estos atributos en cada disparo, asi que el
        cambio se aplica desde el siguiente disparo sin parar la medida.
        No afecta a parametros de hardware (integracion, trigger, etc.),
        que solo se fijan al iniciar (requieren reconfigurar el espectrometro).
        """
        self.do_filter           = do_filter
        self.filter_window       = filter_window
        self.filter_poly         = filter_poly
        self.clip_upper_enabled  = clip_upper_enabled
        self.clip_upper_value    = clip_upper_value
        self.clip_lower_enabled  = clip_lower_enabled
        self.clip_lower_value    = clip_lower_value


# ===============================================
# HILO MAPPING 2D
# ===============================================
class MapThread(QThread):
    progress = pyqtSignal(int)
    done     = pyqtSignal(object)

    def __init__(self, specs, t_ms, wl, selected_wl, grid=5, edge=0):
        super().__init__()
        self.specs       = specs
        self.t_ms        = t_ms
        self.wl          = wl
        self.selected_wl = selected_wl
        self.grid        = grid
        self.edge        = edge

    def run(self):
        idx   = (np.abs(self.wl - self.selected_wl)).argmin()
        cube  = np.zeros((self.grid, self.grid))
        total = self.grid ** 2
        for spec in self.specs:
            spec.prepare_hw(self.t_ms, self.edge)
        k = 0
        for y in range(self.grid):
            for x in range(self.grid):
                threads = [threading.Thread(target=spec.arm) for spec in self.specs]
                for th in threads: th.start()
                for th in threads: th.join()
                s = self.specs[0].read(timeout=60)
                if s is not None:
                    cube[y, x] = float((s - np.min(s))[idx])
                k += 1
                self.progress.emit(int(k / total * 100))
        self.done.emit(cube)


# ===============================================
# PESTANA GENERICA
# ===============================================
class PlotTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(4, 4, 4, 4)
        self.fig = Figure(facecolor="#1a1a2e")
        self.ax  = self.fig.add_subplot(111)
        self._style_ax()
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        layout.addWidget(self.canvas)

    def _style_ax(self):
        self.ax.set_facecolor("#16213e")
        self.ax.tick_params(colors="#aaaacc")
        for sp in self.ax.spines.values(): sp.set_color("#444466")
        self.ax.title.set_color("#ccccff")
        self.ax.xaxis.label.set_color("#aaaacc")
        self.ax.yaxis.label.set_color("#aaaacc")


# ===============================================
# PESTANA MEDIDA
# ===============================================
class MeasureTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._mode       = "view"
        self._zoom_start = None
        self._zoom_patch = None
        self._zoom_cids  = []

        outer = QHBoxLayout(self)
        outer.setContentsMargins(4, 4, 4, 4)
        outer.setSpacing(8)

        left = QWidget()
        lv = QVBoxLayout(left); lv.setContentsMargins(0, 0, 0, 0); lv.setSpacing(4)
        outer.addWidget(left, stretch=1)

        tb = QHBoxLayout(); tb.setSpacing(4)
        self._tb_btns = {}
        for key, label, tip in [
            ("zoom",  "Zoom",  "Arrastra para hacer zoom en una region"),
            ("view",  "Vista", "Solo visionado (sin zoom)"),
            ("reset", "Reset", "Volver a vista original"),
        ]:
            btn = QPushButton(label)
            btn.setToolTip(tip)
            btn.setFixedHeight(24)
            btn.setStyleSheet(self._btn_style(key == "view"))
            if key == "reset":
                btn.clicked.connect(self._reset_zoom)
            else:
                btn.clicked.connect(lambda _c, k=key: self._set_mode(k))
            self._tb_btns[key] = btn
            tb.addWidget(btn)
        tb.addStretch()
        lv.addLayout(tb)

        self.fig = Figure(facecolor="#1a1a2e")
        self.ax  = self.fig.add_subplot(111)
        self._style_ax()
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        lv.addWidget(self.canvas)

        self._connect_zoom_events()

        panel = QWidget(); panel.setFixedWidth(165)
        pl = QVBoxLayout(panel)
        pl.setContentsMargins(4, 4, 4, 4); pl.setSpacing(8)
        outer.addWidget(panel)

        g = QGroupBox("Estado")
        g.setStyleSheet("QGroupBox{color:#ccccff;border:1px solid #444466;"
                        "border-radius:4px;margin-top:6px}"
                        "QGroupBox::title{subcontrol-origin:margin;left:8px}")
        gl = QVBoxLayout(g); gl.setContentsMargins(8, 14, 8, 8); gl.setSpacing(8)

        self.lbl_state  = QLabel("Inactivo")
        self.lbl_state.setStyleSheet("color:#aaaacc;font-size:13px;font-weight:bold")
        self.lbl_state.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_shots  = QLabel("Disparos: 0")
        self.lbl_shots.setStyleSheet("color:#f0a500;font-size:14px;font-weight:bold")
        self.lbl_shots.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_target = QLabel("Objetivo: inf")
        self.lbl_target.setStyleSheet("color:#888899;font-size:11px")
        self.lbl_target.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_time   = QLabel("Ultimo: -")
        self.lbl_time.setStyleSheet("color:#888899;font-size:11px")
        self.lbl_time.setAlignment(Qt.AlignmentFlag.AlignCenter)

        gl.addWidget(self.lbl_state); gl.addWidget(self.lbl_shots)
        gl.addWidget(self.lbl_target); gl.addWidget(self.lbl_time)
        pl.addWidget(g)

        g_pk = QGroupBox("Picos")
        g_pk.setStyleSheet("QGroupBox{color:#ccccff;border:1px solid #444466;"
                           "border-radius:4px;margin-top:6px}"
                           "QGroupBox::title{subcontrol-origin:margin;left:8px}")
        gpkl = QVBoxLayout(g_pk); gpkl.setContentsMargins(8, 14, 8, 8); gpkl.setSpacing(4)
        self.chk_show_peaks = QCheckBox("Mostrar picos")
        self.chk_show_peaks.setChecked(False)
        self.chk_show_peaks.setStyleSheet("color:#aaaacc")
        self.chk_show_peaks.stateChanged.connect(self._on_peaks_toggle)
        gpkl.addWidget(self.chk_show_peaks)
        pl.addWidget(g_pk)

        pl.addStretch()
        self._overlay_lines = []
        self._peak_artists = []
        self._last_wl = None
        self._last_s = None
        self._last_idx = None
        self._last_overlay = False

    # ── toolbar zoom (misma logica que en CompareTab) ────────────────────
    def _btn_style(self, active=False):
        bg = "#3a5f8a" if active else "#2a2a3e"
        return (f"QPushButton{{background:{bg};color:white;border-radius:3px;"
                f"font-size:11px;padding:2px 8px;}}"
                f"QPushButton:hover{{background:#4a6f9a;}}")

    def _set_mode(self, mode):
        self._mode = mode
        for key, btn in self._tb_btns.items():
            if key == "reset": continue
            btn.setStyleSheet(self._btn_style(key == mode))

    def _reset_zoom(self):
        self.ax.autoscale(); self.canvas.draw_idle()

    def _connect_zoom_events(self):
        for cid in self._zoom_cids:
            self.canvas.mpl_disconnect(cid)
        self._zoom_cids = []

        def on_press(event):
            if event.inaxes != self.ax or event.xdata is None: return
            if self._mode != "zoom": return
            self._zoom_start = (event.xdata, event.ydata)
            if self._zoom_patch:
                try: self._zoom_patch.remove()
                except Exception: pass
            ylo, yhi = self.ax.get_ylim()
            self._zoom_patch = MplRect(
                (event.xdata, ylo), 0, yhi - ylo,
                linewidth=1, edgecolor="#f0a500", facecolor="#f0a50022")
            self.ax.add_patch(self._zoom_patch)

        def on_motion(event):
            if event.inaxes != self.ax or event.xdata is None: return
            if self._mode == "zoom" and self._zoom_start:
                x0, _ = self._zoom_start
                if self._zoom_patch:
                    self._zoom_patch.set_width(event.xdata - x0)
                    self.canvas.draw_idle()

        def on_release(event):
            if self._mode == "zoom" and self._zoom_start and event.inaxes == self.ax:
                x0, _ = self._zoom_start
                x1 = event.xdata if event.xdata is not None else x0
                xlo, xhi = min(x0, x1), max(x0, x1)
                if xhi - xlo > 1e-9:
                    self.ax.set_xlim(xlo, xhi)
                    self._autoscale_y_to_xrange(xlo, xhi)
                if self._zoom_patch:
                    try: self._zoom_patch.remove()
                    except Exception: pass
                    self._zoom_patch = None
                self._zoom_start = None
                self.canvas.draw_idle()

        self._zoom_cids.append(self.canvas.mpl_connect("button_press_event",   on_press))
        self._zoom_cids.append(self.canvas.mpl_connect("motion_notify_event",  on_motion))
        self._zoom_cids.append(self.canvas.mpl_connect("button_release_event", on_release))

    def _autoscale_y_to_xrange(self, xlo, xhi, margin=0.08):
        ymin, ymax = None, None
        for line in self.ax.get_lines():
            xdata = line.get_xdata()
            ydata = line.get_ydata()
            if len(xdata) == 0:
                continue
            mask = (np.asarray(xdata) >= xlo) & (np.asarray(xdata) <= xhi)
            if not np.any(mask):
                continue
            y_sel = np.asarray(ydata)[mask]
            y_sel = y_sel[np.isfinite(y_sel)]
            if y_sel.size == 0:
                continue
            lo, hi = float(y_sel.min()), float(y_sel.max())
            ymin = lo if ymin is None else min(ymin, lo)
            ymax = hi if ymax is None else max(ymax, hi)
        if ymin is None or ymax is None:
            return
        span = ymax - ymin
        if span <= 1e-12:
            span = abs(ymax) if ymax != 0 else 1.0
        pad = span * margin
        self.ax.set_ylim(ymin - pad, ymax + pad)

    def _on_peaks_toggle(self, _state):
        # redibuja el ultimo disparo con/sin marcadores de picos
        if self._last_wl is not None and self._last_s is not None:
            self.update_shot(self._last_idx, self._last_wl, self._last_s,
                              overlay=self._last_overlay)

    def _style_ax(self):
        self.ax.set_facecolor("#16213e")
        self.ax.tick_params(colors="#aaaacc")
        for sp in self.ax.spines.values(): sp.set_color("#444466")
        self.ax.title.set_color("#ccccff")
        self.ax.xaxis.label.set_color("#aaaacc")
        self.ax.yaxis.label.set_color("#aaaacc")

    def set_state(self, text, color="#aaaacc"):
        self.lbl_state.setText(text)
        self.lbl_state.setStyleSheet(f"color:{color};font-size:13px;font-weight:bold")

    def set_target(self, n):
        self.lbl_target.setText(f"Objetivo: {'inf' if n == 0 else n}")

    def _draw_peaks(self, wl, s, color):
        for art in self._peak_artists:
            try: art.remove()
            except Exception: pass
        self._peak_artists = []
        if not self.chk_show_peaks.isChecked():
            return
        thresh = 0.08 * (s.max() - s.min()) if s.max() > s.min() else 0
        peaks, _ = find_peaks(s, prominence=thresh, distance=10)
        if len(peaks) == 0:
            return
        marks, = self.ax.plot(wl[peaks], s[peaks], "v", color=color,
                               markersize=6, markeredgecolor="#ffffff44", zorder=5)
        self._peak_artists.append(marks)
        for p in peaks:
            ann = self.ax.annotate(f"{wl[p]:.1f}", xy=(wl[p], s[p]),
                                    xytext=(0, 8), textcoords="offset points",
                                    ha="center", fontsize=7, color=color)
            self._peak_artists.append(ann)

    def update_shot(self, idx, wl, s, overlay=False, max_overlay=10):
        self._last_wl, self._last_s = wl, s
        self._last_idx, self._last_overlay = idx, overlay

        # conservar limites actuales si el usuario ha hecho zoom manualmente
        keep_xlim = self.ax.get_xlim() if (self._last_idx not in (None, 1)) else None
        keep_ylim = self.ax.get_ylim() if (self._last_idx not in (None, 1)) else None
        had_zoom = self._mode == "zoom" or keep_xlim not in (None, (0.0, 1.0))

        self.lbl_shots.setText(f"Disparos: {idx}")
        self.lbl_time.setText(f"Ultimo: {time.strftime('%H:%M:%S')}")
        if overlay:
            alphas = np.linspace(0.12, 1.0, max_overlay)
            color  = COLORS[(idx - 1) % len(COLORS)]
            line,  = self.ax.plot(wl, s, color=color, linewidth=1, alpha=1.0)
            self._overlay_lines.append(line)
            if len(self._overlay_lines) > max_overlay:
                self._overlay_lines.pop(0).remove()
            for i, ln in enumerate(self._overlay_lines):
                ln.set_alpha(alphas[i + max_overlay - len(self._overlay_lines)])
            self._draw_peaks(wl, s, color)
        else:
            xlim_before = self.ax.get_xlim()
            ylim_before = self.ax.get_ylim()
            has_data_before = len(self.ax.get_lines()) > 0
            self.ax.clear(); self._style_ax()
            self._peak_artists = []
            color = "#50fa7b"
            self.ax.plot(wl, s, color=color, linewidth=1.2)
            self._draw_peaks(wl, s, color)
            if has_data_before and self._mode == "zoom":
                self.ax.set_xlim(xlim_before)
                self.ax.set_ylim(ylim_before)
        self.ax.set_xlabel("Longitud de onda (nm)")
        self.ax.set_ylabel("Intensidad norm.")
        self.ax.set_title(f"Disparo #{idx}")
        self.canvas.draw_idle()

    def reset(self):
        self._overlay_lines = []
        self._peak_artists = []
        self._last_wl = None
        self._last_s = None
        self.ax.clear(); self._style_ax()
        self.canvas.draw()
        self.lbl_shots.setText("Disparos: 0")
        self.lbl_time.setText("Ultimo: -")

class CompareTab(QWidget):
    # señal interna: el usuario pulsó Añadir — la maneja App
    add_sim_requested = pyqtSignal()
    delete_selected_requested = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._mode       = "view"
        self._zoom_start = None
        self._zoom_patch = None
        self._zoom_cids  = []

        outer = QHBoxLayout(self)
        outer.setContentsMargins(4, 4, 4, 4); outer.setSpacing(6)

        # ── columna izquierda: toolbar + canvas ──────────────────────────
        left = QWidget()
        lv   = QVBoxLayout(left); lv.setContentsMargins(0, 0, 0, 0); lv.setSpacing(4)
        outer.addWidget(left, stretch=1)

        tb = QHBoxLayout(); tb.setSpacing(4)
        self._tb_btns = {}
        for key, label, tip in [
            ("cursor_r", "Cursor R",  "Click fija cursor rojo (lambda_R)"),
            ("cursor_b", "Cursor B",  "Click fija cursor azul (lambda_B)"),
            ("zoom",     "Zoom",      "Arrastra para hacer zoom en region"),
            ("view",     "Vista",     "Solo visionado"),
            ("reset",    "Reset",     "Volver a vista original"),
        ]:
            btn = QPushButton(label)
            btn.setToolTip(tip)
            btn.setFixedHeight(26)
            btn.setStyleSheet(self._btn_style(key == "view"))
            if key == "reset":
                btn.clicked.connect(self._reset_zoom)
            else:
                btn.clicked.connect(lambda _c, k=key: self._set_mode(k))
            self._tb_btns[key] = btn
            tb.addWidget(btn)
        tb.addStretch()
        lv.addLayout(tb)

        self.fig = Figure(facecolor="#1a1a2e")
        self.ax  = self.fig.add_subplot(111)
        self._style_ax()
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        lv.addWidget(self.canvas)

        # ── columna derecha ───────────────────────────────────────────────
        right = QWidget(); right.setFixedWidth(215)
        rl = QVBoxLayout(right); rl.setContentsMargins(4, 4, 4, 4); rl.setSpacing(4)

        right_scroll = QScrollArea()
        right_scroll.setWidget(right)
        right_scroll.setWidgetResizable(True)
        right_scroll.setFixedWidth(230)
        right_scroll.setFrameShape(QFrame.Shape.NoFrame)
        right_scroll.setStyleSheet(
            "QScrollArea{background:transparent;border:none;}"
            "QScrollArea > QWidget > QWidget{background:transparent;}")
        outer.addWidget(right_scroll)

        # --- 1. Lista de gráficas (espectros cargados) ---
        g_spec = QGroupBox("Graficas")
        g_spec.setStyleSheet("QGroupBox{color:#ccccff;border:1px solid #444466;"
                             "border-radius:4px;margin-top:6px}"
                             "QGroupBox::title{subcontrol-origin:margin;left:8px}")
        gsl = QVBoxLayout(g_spec); gsl.setContentsMargins(4, 12, 4, 4); gsl.setSpacing(2)

        self.chk_normalize = QCheckBox("Normalizar (0-1)")
        self.chk_normalize.setChecked(False)
        self.chk_normalize.setStyleSheet("color:#f0a500;font-size:10px;font-weight:bold")
        self.chk_normalize.setToolTip(
            "Escala cada espectro visible a su propio maximo (0-1) solo para "
            "esta vista. Util para comparar formas cuando las medidas reales "
            "y las simulaciones NIST tienen rangos de intensidad muy distintos. "
            "No afecta a los datos guardados ni a Depth Profile/exportacion.")
        gsl.addWidget(self.chk_normalize)

        sel_row = QHBoxLayout(); sel_row.setSpacing(4)
        sel_hint = QLabel("Sel.")
        sel_hint.setStyleSheet("color:#888899;font-size:9px")
        sel_row.addWidget(sel_hint)
        sel_row.addStretch()
        self.btn_select_all = QPushButton("Todos")
        self.btn_select_all.setFixedHeight(18)
        self.btn_select_all.setStyleSheet(
            "QPushButton{background:#2a2a3e;color:#aaaacc;font-size:9px;border-radius:2px;padding:0 6px;}"
            "QPushButton:hover{background:#3a3a5e;}")
        self.btn_select_all.clicked.connect(self._select_all_for_delete)
        sel_row.addWidget(self.btn_select_all)
        self.btn_select_none = QPushButton("Ninguno")
        self.btn_select_none.setFixedHeight(18)
        self.btn_select_none.setStyleSheet(
            "QPushButton{background:#2a2a3e;color:#aaaacc;font-size:9px;border-radius:2px;padding:0 6px;}"
            "QPushButton:hover{background:#3a3a5e;}")
        self.btn_select_none.clicked.connect(self._select_none_for_delete)
        sel_row.addWidget(self.btn_select_none)
        gsl.addLayout(sel_row)

        scroll_spec = QScrollArea(); scroll_spec.setWidgetResizable(True)
        scroll_spec.setFrameShape(QFrame.Shape.NoFrame)
        scroll_spec.setStyleSheet("background:transparent")
        self._cb_container = QWidget()
        self._cb_layout    = QVBoxLayout(self._cb_container)
        self._cb_layout.setContentsMargins(0, 0, 0, 0); self._cb_layout.setSpacing(2)
        self._cb_layout.addStretch()
        scroll_spec.setWidget(self._cb_container)
        gsl.addWidget(scroll_spec)

        self.btn_delete_selected = QPushButton("Eliminar seleccionados")
        self.btn_delete_selected.setFixedHeight(24)
        self.btn_delete_selected.setStyleSheet(
            "QPushButton{background:#7a2a2a;color:white;border-radius:3px;font-size:10px;}"
            "QPushButton:hover{background:#aa3333;}")
        self.btn_delete_selected.clicked.connect(self.delete_selected_requested)
        gsl.addWidget(self.btn_delete_selected)

        make_collapsible(g_spec)
        rl.addWidget(g_spec, stretch=2)

        # --- 1b. Filtrado posterior (solo vista, no altera datos guardados) ---
        g_filt = QGroupBox("Filtrado (vista)")
        g_filt.setStyleSheet("QGroupBox{color:#ccccff;border:1px solid #444466;"
                             "border-radius:4px;margin-top:6px}"
                             "QGroupBox::title{subcontrol-origin:margin;left:8px}")
        gfl = QVBoxLayout(g_filt); gfl.setContentsMargins(6, 12, 6, 6); gfl.setSpacing(3)
        gfl_hint = QLabel("Se aplica solo a esta vista;\nno afecta a datos guardados/exportados.")
        gfl_hint.setStyleSheet("color:#666688;font-size:8px")
        gfl.addWidget(gfl_hint)

        self.chk_post_sg = QCheckBox("Savitzky-Golay")
        self.chk_post_sg.setChecked(False)
        self.chk_post_sg.setStyleSheet("color:#aaaacc;font-size:10px")
        gfl.addWidget(self.chk_post_sg)

        sg_row = QHBoxLayout(); sg_row.setSpacing(4)
        sg_lbl1 = QLabel("Vent.:"); sg_lbl1.setStyleSheet("color:#888899;font-size:9px")
        sg_row.addWidget(sg_lbl1)
        self.spin_post_sg_window = QSpinBox()
        self.spin_post_sg_window.setRange(3, 99)
        self.spin_post_sg_window.setSingleStep(2)
        self.spin_post_sg_window.setValue(11)
        self.spin_post_sg_window.setFixedWidth(52)
        self.spin_post_sg_window.valueChanged.connect(self._on_post_sg_window_changed)
        sg_row.addWidget(self.spin_post_sg_window)
        sg_lbl2 = QLabel("Ord.:"); sg_lbl2.setStyleSheet("color:#888899;font-size:9px")
        sg_row.addWidget(sg_lbl2)
        self.spin_post_sg_poly = QSpinBox()
        self.spin_post_sg_poly.setRange(1, 7)
        self.spin_post_sg_poly.setValue(3)
        self.spin_post_sg_poly.setFixedWidth(44)
        sg_row.addWidget(self.spin_post_sg_poly)
        gfl.addLayout(sg_row)

        sep_pf = QFrame(); sep_pf.setFrameShape(QFrame.Shape.HLine); sep_pf.setStyleSheet("color:#444466")
        gfl.addWidget(sep_pf)

        self.chk_post_clip_upper = QCheckBox("Limitar superiores")
        self.chk_post_clip_upper.setChecked(False)
        self.chk_post_clip_upper.setStyleSheet("color:#aaaacc;font-size:10px")
        gfl.addWidget(self.chk_post_clip_upper)
        cu_row = QHBoxLayout(); cu_row.setSpacing(4)
        cu_lbl = QLabel("Umbral max:"); cu_lbl.setStyleSheet("color:#888899;font-size:9px")
        cu_row.addWidget(cu_lbl)
        self.spin_post_clip_upper = QDoubleSpinBox()
        self.spin_post_clip_upper.setRange(0.0, 1e9)
        self.spin_post_clip_upper.setDecimals(3)
        self.spin_post_clip_upper.setValue(1.0)
        self.spin_post_clip_upper.setSingleStep(0.1)
        self.spin_post_clip_upper.setFixedWidth(68)
        cu_row.addWidget(self.spin_post_clip_upper)
        gfl.addLayout(cu_row)

        self.chk_post_clip_lower = QCheckBox("Limitar inferiores")
        self.chk_post_clip_lower.setChecked(False)
        self.chk_post_clip_lower.setStyleSheet("color:#aaaacc;font-size:10px")
        gfl.addWidget(self.chk_post_clip_lower)
        cl_row = QHBoxLayout(); cl_row.setSpacing(4)
        cl_lbl = QLabel("Umbral min:"); cl_lbl.setStyleSheet("color:#888899;font-size:9px")
        cl_row.addWidget(cl_lbl)
        self.spin_post_clip_lower = QDoubleSpinBox()
        self.spin_post_clip_lower.setRange(-1e9, 1e9)
        self.spin_post_clip_lower.setDecimals(3)
        self.spin_post_clip_lower.setValue(0.0)
        self.spin_post_clip_lower.setSingleStep(0.1)
        self.spin_post_clip_lower.setFixedWidth(68)
        cl_row.addWidget(self.spin_post_clip_lower)
        gfl.addLayout(cl_row)

        make_collapsible(g_filt, start_expanded=False)
        rl.addWidget(g_filt)

        # --- 2. Cursores ---
        g_cur = QGroupBox("Cursores")
        g_cur.setStyleSheet("QGroupBox{color:#ccccff;border:1px solid #444466;"
                            "border-radius:4px;margin-top:6px}"
                            "QGroupBox::title{subcontrol-origin:margin;left:8px}")
        gcl = QVBoxLayout(g_cur); gcl.setContentsMargins(8, 14, 8, 6); gcl.setSpacing(4)
        hint = QLabel("Cursor R = click izq.\nCursor B = click der.")
        hint.setStyleSheet("color:#888899;font-size:9px")
        hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        gcl.addWidget(hint)
        sep = QFrame(); sep.setFrameShape(QFrame.Shape.HLine); sep.setStyleSheet("color:#444466")
        gcl.addWidget(sep)
        self.lbl_red   = self._clabel("lambda_R:", "#ff5555")
        self.lbl_blue  = self._clabel("lambda_B:", "#8be9fd")
        self.lbl_delta = self._clabel("Delta-wl:", "#f0a500", bold=True)
        gcl.addWidget(self.lbl_red); gcl.addWidget(self.lbl_blue)
        sep2 = QFrame(); sep2.setFrameShape(QFrame.Shape.HLine); sep2.setStyleSheet("color:#444466")
        gcl.addWidget(sep2); gcl.addWidget(self.lbl_delta)
        make_collapsible(g_cur)
        rl.addWidget(g_cur)

        # --- 3. Lista de elementos NIST ---
        g_elem = QGroupBox("Elementos NIST")
        g_elem.setStyleSheet("QGroupBox{color:#aaaacc;border:1px solid #444466;"
                             "border-radius:4px;margin-top:6px}"
                             "QGroupBox::title{subcontrol-origin:margin;left:8px}")
        gel = QVBoxLayout(g_elem); gel.setContentsMargins(4, 12, 4, 4); gel.setSpacing(2)
        self.elem_list = QListWidget()
        self.elem_list.setStyleSheet(
            "QListWidget{background:#16213e;border:none;}"
            "QListWidget::item{color:#ccccff;padding:2px 4px;font-size:11px;}"
            "QListWidget::item:selected{background:#3a5f8a;color:white;}"
            "QListWidget::item:hover{background:#2a2a5e;}"
        )
        for el in ELEMENTS:
            self.elem_list.addItem(el)
        self.elem_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        gel.addWidget(self.elem_list)
        make_collapsible(g_elem)
        rl.addWidget(g_elem, stretch=2)

        # --- 4. Parametros del plasma ---
        g_par = QGroupBox("Parametros plasma")
        g_par.setStyleSheet("QGroupBox{color:#aaaacc;border:1px solid #444466;"
                            "border-radius:4px;margin-top:6px}"
                            "QGroupBox::title{subcontrol-origin:margin;left:8px}")
        gpl = QVBoxLayout(g_par); gpl.setContentsMargins(6, 14, 6, 6); gpl.setSpacing(3)

        def prow(lbl_txt, widget):
            r = QHBoxLayout(); r.setSpacing(4)
            l = QLabel(lbl_txt); l.setStyleSheet("color:#888899;font-size:9px;")
            l.setFixedWidth(76)
            r.addWidget(l); r.addWidget(widget)
            return r

        self.sim_Te = QDoubleSpinBox()
        self.sim_Te.setRange(0.5, 5.0); self.sim_Te.setValue(1.0)
        self.sim_Te.setSingleStep(0.1); self.sim_Te.setDecimals(2); self.sim_Te.setFixedWidth(68)

        self.sim_Ne = QDoubleSpinBox()
        self.sim_Ne.setRange(0.7, 1.3); self.sim_Ne.setValue(1.0)
        self.sim_Ne.setSingleStep(0.05); self.sim_Ne.setDecimals(2); self.sim_Ne.setFixedWidth(68)
        self.sim_Ne.setToolTip("x10^17 cm^-3")

        self.sim_pct = QSpinBox()
        self.sim_pct.setRange(500, 1000); self.sim_pct.setValue(1000)
        self.sim_pct.setFixedWidth(68)
        self.sim_pct.setToolTip("Porcentaje del elemento seleccionado")

        self.sim_res = QSpinBox()
        self.sim_res.setRange(1, 100); self.sim_res.setValue(1000)
        self.sim_res.setSingleStep(100); self.sim_res.setFixedWidth(68)

        self.sim_low = QSpinBox()
        self.sim_low.setRange(100, 900); self.sim_low.setValue(200)
        self.sim_low.setSingleStep(50); self.sim_low.setFixedWidth(68)

        self.sim_up = QSpinBox()
        self.sim_up.setRange(200, 1100); self.sim_up.setValue(1000)
        self.sim_up.setSingleStep(50); self.sim_up.setFixedWidth(68)

        self.sim_ion = QSpinBox()
        self.sim_ion.setRange(1, 5); self.sim_ion.setValue(3); self.sim_ion.setFixedWidth(68)

        self.sim_ws = QComboBox()
        self.sim_ws.addItems(["static", "dynamic"]); self.sim_ws.setFixedWidth(68)
        self.sim_ws.setStyleSheet("QComboBox{color:black;background:#e8e8e8;font-size:10px;}")
        self.sim_ws.setToolTip("static=cache local  dynamic=NIST en vivo")

        gpl.addLayout(prow("Te (eV):",      self.sim_Te))
        gpl.addLayout(prow("Ne (x10^17):",  self.sim_Ne))
        gpl.addLayout(prow("Porcentaje %:", self.sim_pct))
        gpl.addLayout(prow("Resolucion:",   self.sim_res))
        gpl.addLayout(prow("wl min (nm):",  self.sim_low))
        gpl.addLayout(prow("wl max (nm):",  self.sim_up))
        gpl.addLayout(prow("Max ion:",      self.sim_ion))
        gpl.addLayout(prow("Fuente NIST:",  self.sim_ws))
        make_collapsible(g_par)
        rl.addWidget(g_par)

        # --- 5. Boton Añadir ---
        self.btn_add_sim = QPushButton("Añadir simulacion")
        self.btn_add_sim.setFixedHeight(30)
        self.btn_add_sim.setStyleSheet(
            "QPushButton{background:#2a4a7a;color:white;border-radius:4px;font-weight:bold;}"
            "QPushButton:hover{background:#3a6a9a;}"
        )
        self.btn_add_sim.clicked.connect(self.add_sim_requested)
        rl.addWidget(self.btn_add_sim)

        # estado interno
        self.checkboxes = []
        self._cursor_r = self._cursor_b = None
        self._wl_r = self._wl_b = None

    # ── toolbar helpers ──────────────────────────────────────────────────
    def _btn_style(self, active=False):
        bg = "#3a5f8a" if active else "#2a2a3e"
        return (f"QPushButton{{background:{bg};color:white;border-radius:3px;"
                f"font-size:11px;padding:2px 8px;}}"
                f"QPushButton:hover{{background:#4a6f9a;}}")

    def _set_mode(self, mode):
        self._mode = mode
        for key, btn in self._tb_btns.items():
            if key == "reset": continue
            btn.setStyleSheet(self._btn_style(key == mode))

    def _reset_zoom(self):
        self.ax.autoscale(); self.canvas.draw_idle()

    def _on_post_sg_window_changed(self, val):
        if val % 2 == 0:
            self.spin_post_sg_window.blockSignals(True)
            self.spin_post_sg_window.setValue(val + 1)
            self.spin_post_sg_window.blockSignals(False)

    def _clabel(self, prefix, color, bold=False):
        lbl = QLabel(f"{prefix}  -")
        lbl.setStyleSheet(f"color:{color};font-size:11px;"
                          f"font-weight:{'bold' if bold else 'normal'}")
        return lbl

    def _style_ax(self):
        self.ax.set_facecolor("#16213e")
        self.ax.tick_params(colors="#aaaacc")
        for sp in self.ax.spines.values(): sp.set_color("#444466")
        self.ax.title.set_color("#ccccff")
        self.ax.xaxis.label.set_color("#aaaacc")
        self.ax.yaxis.label.set_color("#aaaacc")

    # ── selección múltiple para borrado ──────────────────────────────────
    def _select_all_for_delete(self):
        for _, _, _, del_cb in self.checkboxes:
            del_cb.setChecked(True)

    def _select_none_for_delete(self):
        for _, _, _, del_cb in self.checkboxes:
            del_cb.setChecked(False)

    def selected_for_delete_indices(self):
        return [i for i, (_, _, _, del_cb) in enumerate(self.checkboxes) if del_cb.isChecked()]

    # ── gestión de filas ─────────────────────────────────────────────────
    def add_spectrum_row(self, display_name, session_label, color, on_change, on_delete):
        row = QWidget()
        rl  = QHBoxLayout(row); rl.setContentsMargins(0, 0, 0, 0); rl.setSpacing(2)

        del_cb = QCheckBox()
        del_cb.setToolTip("Marcar para borrado multiple")
        del_cb.setStyleSheet("QCheckBox::indicator{width:12px;height:12px}")
        del_cb.setChecked(False)
        rl.addWidget(del_cb)

        cb = QCheckBox(f"  {display_name}")
        cb.setToolTip(f"Conjunto: {session_label}")
        cb.setStyleSheet(f"QCheckBox{{color:{color};font-size:11px}}"
                         "QCheckBox::indicator{width:12px;height:12px}")
        cb.setChecked(True)
        cb.stateChanged.connect(on_change)
        rl.addWidget(cb, stretch=1)

        set_lbl = QLabel(session_label[:6])
        set_lbl.setStyleSheet("color:#555577;font-size:8px;")
        set_lbl.setFixedWidth(36)
        rl.addWidget(set_lbl)

        del_btn = QPushButton("X")
        del_btn.setFixedSize(16, 16)
        del_btn.setStyleSheet(
            "QPushButton{background:#7a2a2a;color:white;border-radius:2px;"
            "font-size:8px;padding:0;}"
            "QPushButton:hover{background:#aa3333;}"
        )
        del_btn.clicked.connect(on_delete)
        rl.addWidget(del_btn)

        item = self._cb_layout.takeAt(self._cb_layout.count() - 1)
        self._cb_layout.addWidget(row)
        self._cb_layout.addItem(item)
        self.checkboxes.append((cb, row, set_lbl, del_cb))

    def update_row_name(self, idx, display_name, color):
        cb, _, _, _ = self.checkboxes[idx]
        cb.setText(f"  {display_name}")
        cb.setStyleSheet(f"QCheckBox{{color:{color};font-size:11px}}"
                         "QCheckBox::indicator{width:12px;height:12px}")

    def visible_indices(self):
        return [i for i, (cb, _, _, _) in enumerate(self.checkboxes) if cb.isChecked()]

    def clear_spectra(self):
        for _, row, _, _ in self.checkboxes:
            row.deleteLater()
        self.checkboxes = []
        while self._cb_layout.count() > 1:
            item = self._cb_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self._wl_r = self._wl_b = None
        self._cursor_r = self._cursor_b = None
        self._zoom_start = None
        self._zoom_patch = None
        self.ax.clear(); self._style_ax()
        self.canvas.draw()

    def update_cursor_display(self):
        r = f"{self._wl_r:.2f} nm" if self._wl_r is not None else "-"
        b = f"{self._wl_b:.2f} nm" if self._wl_b is not None else "-"
        self.lbl_red.setText(f"lambda_R:  {r}")
        self.lbl_blue.setText(f"lambda_B:  {b}")
        if self._wl_r is not None and self._wl_b is not None:
            d = self._wl_r - self._wl_b
            self.lbl_delta.setText(f"Delta-wl:  {'+' if d >= 0 else ''}{d:.2f} nm")
        elif self._wl_r is not None:
            self.lbl_delta.setText(f"Delta-wl:  {self._wl_r:.2f} nm")
        else:
            self.lbl_delta.setText("Delta-wl:  -")

# ===============================================
# APP PRINCIPAL
# ===============================================
class App(QMainWindow):
    def __init__(self):
        super().__init__()
        self.specs          = []
        self.wl             = None
        self.last_spectrum  = None
        self.saved          = []
        self.selected_wl    = None
        self._trig_thread   = None
        self._map_thread    = None
        self._session_count = 0
        self._peak_cache    = {}

        self.setWindowTitle("LIBS PRO - PyQt6")
        self.resize(1500, 860)
        self._build_ui()

    def log(self, msg):
        print(msg)
        self.statusBar().showMessage(msg)

    # ── conexión ────────────────────────────────────────────────────────
    def connect_devices(self):
        n = avaspec.AVS_GetNrOfDevices()
        if n <= 0: self.log("No hay dispositivos"); return
        ids = (avaspec.AvsIdentityType * n)()
        req = ctypes.c_int()
        avaspec.AVS_GetList(ctypes.sizeof(ids), ctypes.byref(req), ids)
        self.specs = []
        for i in range(n):
            spec = Spectrometer(self.log, ids[i])
            if spec.connect(): self.specs.append(spec)
        if self.specs:
            self.log(f"{len(self.specs)} canal(es) conectados")
            self.btn_connect.setStyleSheet(
                "QPushButton{background:#1a6b3a;color:white;border-radius:4px;font-weight:bold;}"
                "QPushButton:hover{background:#256b3a;}")
        else:
            self.log("Error al conectar")
            self.btn_connect.setStyleSheet("")

    def disconnect_devices(self):
        self.stop()
        for s in self.specs: s.disconnect()
        self.specs = []
        self.btn_connect.setStyleSheet("")
        self.log("Desconectado")

    # ── UI ───────────────────────────────────────────────────────────────
    def _build_ui(self):
        root = QWidget(); self.setCentralWidget(root)
        rl = QHBoxLayout(root); rl.setSpacing(10)

        left = QWidget(); left.setFixedWidth(250)
        ll = QVBoxLayout(left); ll.setSpacing(8); ll.setContentsMargins(6, 6, 6, 6)

        left_scroll = QScrollArea()
        left_scroll.setWidget(left)
        left_scroll.setWidgetResizable(True)
        left_scroll.setFixedWidth(266)
        left_scroll.setFrameShape(QFrame.Shape.NoFrame)
        left_scroll.setStyleSheet(
            "QScrollArea{background:transparent;border:none;}"
            "QScrollArea > QWidget > QWidget{background:transparent;}")
        rl.addWidget(left_scroll)

        g1 = QGroupBox("Adquisicion")
        v1 = QVBoxLayout(g1)

        v1.addWidget(QLabel("Integracion"))
        int_row = QHBoxLayout(); int_row.setSpacing(4)
        self.t = QLineEdit("50")
        int_row.addWidget(self.t)
        self.cmb_t_unit = QComboBox()
        self.cmb_t_unit.addItems(["ms", "us", "ns"])
        self.cmb_t_unit.setFixedWidth(55)
        self.cmb_t_unit.setStyleSheet(
            "QComboBox{color:black;background:#e8e8e8;border:1px solid #aaa;"
            "border-radius:3px;padding:1px 4px;}"
            "QComboBox QAbstractItemView{color:black;background:#f0f0f0;}")
        int_row.addWidget(self.cmb_t_unit)
        v1.addLayout(int_row)

        v1.addWidget(QLabel("Tiempo de retardo"))
        delay_row = QHBoxLayout(); delay_row.setSpacing(4)
        self.t_re = QLineEdit("0")
        delay_row.addWidget(self.t_re)
        self.cmb_delay_unit = QComboBox()
        self.cmb_delay_unit.addItems(["us", "ms", "ns"])
        self.cmb_delay_unit.setFixedWidth(55)
        self.cmb_delay_unit.setStyleSheet(
            "QComboBox{color:black;background:#e8e8e8;border:1px solid #aaa;"
            "border-radius:3px;padding:1px 4px;}"
            "QComboBox QAbstractItemView{color:black;background:#f0f0f0;}")
        delay_row.addWidget(self.cmb_delay_unit)
        v1.addLayout(delay_row)

        v1.addWidget(QLabel("Trigger:"))
        self.cmb_edge = QComboBox()
        self.cmb_edge.addItems(["Interno", "Externo"])
        self.cmb_edge.setStyleSheet(
            "QComboBox{color:black;background:#e8e8e8;border:1px solid #aaa;"
            "border-radius:3px;padding:2px 6px;}"
            "QComboBox QAbstractItemView{color:black;background:#f0f0f0;}")
        v1.addWidget(self.cmb_edge)
        v1.addWidget(QLabel("N disparos (0 = inf):"))
        self.spin_shots = QLineEdit("0"); v1.addWidget(self.spin_shots)
        self.chk_overlay = QCheckBox("Overlay de disparos")
        self.chk_overlay.setChecked(False)
        self.chk_overlay.setStyleSheet("color:#aaaacc"); v1.addWidget(self.chk_overlay)
        v1.addWidget(QLabel("Acumulacion:"))
        self.cmb_accum = QComboBox()
        self.cmb_accum.addItems(["Sin acumular", "Suma acumulada", "Promedio"])
        self.cmb_accum.setStyleSheet(
            "QComboBox{color:black;background:#e8e8e8;border:1px solid #aaa;"
            "border-radius:3px;padding:2px 6px;}"
            "QComboBox QAbstractItemView{color:black;background:#f0f0f0;}")
        v1.addWidget(self.cmb_accum)
        make_collapsible(g1)
        ll.addWidget(g1)

        # ── grupo separado: filtros (menos denso, mas legible) ───────────
        g_filt = QGroupBox("Filtros")
        vf = QVBoxLayout(g_filt)

        self.chk_filter = QCheckBox("Filtro Savitzky-Golay")
        self.chk_filter.setChecked(False)
        self.chk_filter.setStyleSheet("color:#aaaacc")
        self.chk_filter.stateChanged.connect(self._on_filter_toggle)
        vf.addWidget(self.chk_filter)

        filt_row = QHBoxLayout(); filt_row.setSpacing(4)
        lbl_win = QLabel("Ventana:"); lbl_win.setStyleSheet("color:#888899;font-size:10px")
        filt_row.addWidget(lbl_win)
        self.spin_filter_window = QSpinBox()
        self.spin_filter_window.setRange(3, 99)
        self.spin_filter_window.setSingleStep(2)
        self.spin_filter_window.setValue(11)
        self.spin_filter_window.valueChanged.connect(self._on_filter_window_changed)
        filt_row.addWidget(self.spin_filter_window)
        vf.addLayout(filt_row)

        poly_row = QHBoxLayout(); poly_row.setSpacing(4)
        lbl_poly = QLabel("Orden:"); lbl_poly.setStyleSheet("color:#888899;font-size:10px")
        poly_row.addWidget(lbl_poly)
        self.spin_filter_poly = QSpinBox()
        self.spin_filter_poly.setRange(1, 7)
        self.spin_filter_poly.setValue(3)
        poly_row.addWidget(self.spin_filter_poly)
        vf.addLayout(poly_row)

        sep_clip = QFrame(); sep_clip.setFrameShape(QFrame.Shape.HLine)
        sep_clip.setStyleSheet("color:#444466"); vf.addWidget(sep_clip)

        vf.addWidget(QLabel("Filtros de umbral (recorte):"))

        self.chk_clip_upper = QCheckBox("Limitar superiores")
        self.chk_clip_upper.setChecked(False)
        self.chk_clip_upper.setStyleSheet("color:#aaaacc")
        self.chk_clip_upper.stateChanged.connect(self._on_clip_toggle)
        vf.addWidget(self.chk_clip_upper)

        clip_up_row = QHBoxLayout(); clip_up_row.setSpacing(4)
        lbl_clip_up = QLabel("Umbral max:"); lbl_clip_up.setStyleSheet("color:#888899;font-size:10px")
        clip_up_row.addWidget(lbl_clip_up)
        self.spin_clip_upper = QDoubleSpinBox()
        self.spin_clip_upper.setRange(0.0, 1e9)
        self.spin_clip_upper.setDecimals(3)
        self.spin_clip_upper.setValue(1.0)
        self.spin_clip_upper.setSingleStep(0.1)
        self.spin_clip_upper.setEnabled(False)
        self.spin_clip_upper.setToolTip(
            "Cualquier valor de intensidad por encima de este umbral se "
            "recorta (satura) a este mismo valor.")
        clip_up_row.addWidget(self.spin_clip_upper)
        vf.addLayout(clip_up_row)

        self.chk_clip_lower = QCheckBox("Limitar inferiores")
        self.chk_clip_lower.setChecked(False)
        self.chk_clip_lower.setStyleSheet("color:#aaaacc")
        self.chk_clip_lower.stateChanged.connect(self._on_clip_toggle)
        vf.addWidget(self.chk_clip_lower)

        clip_lo_row = QHBoxLayout(); clip_lo_row.setSpacing(4)
        lbl_clip_lo = QLabel("Umbral min:"); lbl_clip_lo.setStyleSheet("color:#888899;font-size:10px")
        clip_lo_row.addWidget(lbl_clip_lo)
        self.spin_clip_lower = QDoubleSpinBox()
        self.spin_clip_lower.setRange(-1e9, 1e9)
        self.spin_clip_lower.setDecimals(3)
        self.spin_clip_lower.setValue(0.0)
        self.spin_clip_lower.setSingleStep(0.1)
        self.spin_clip_lower.setEnabled(False)
        self.spin_clip_lower.setToolTip(
            "Cualquier valor de intensidad por debajo de este umbral se "
            "recorta (satura) a este mismo valor.")
        clip_lo_row.addWidget(self.spin_clip_lower)
        vf.addLayout(clip_lo_row)
        make_collapsible(g_filt)
        ll.addWidget(g_filt)

        # propagar cualquier cambio de filtro EN CALIENTE al hilo de
        # adquisicion si hay una medida en curso (sin necesidad de parar
        # y volver a iniciar)
        for w in (self.chk_filter, self.chk_clip_upper, self.chk_clip_lower):
            w.stateChanged.connect(self._push_live_filter_params)
        for w in (self.spin_filter_window, self.spin_filter_poly,
                  self.spin_clip_upper, self.spin_clip_lower):
            w.valueChanged.connect(self._push_live_filter_params)

        g2 = QGroupBox("Conexion")
        v2 = QVBoxLayout(g2)
        self.btn_connect=self._btn(v2, "Conectar",    self.connect_devices)
        self.btn_disconnect = self._btn(v2, "Desconectar", self.disconnect_devices)
        make_collapsible(g2)
        ll.addWidget(g2)

        g3 = QGroupBox("Medida (trigger)")
        g3.setStyleSheet("QGroupBox{color:#50fa7b;font-weight:bold;"
                         "border:1px solid #50fa7b55;border-radius:4px;margin-top:6px}"
                         "QGroupBox::title{subcontrol-origin:margin;left:8px}")
        v3 = QVBoxLayout(g3)
        self._btn(v3, "Iniciar medida", self.start, "#1a6b3a")
        self._btn(v3, "Parar",          self.stop,  "#7a2a2a")
        make_collapsible(g3)
        ll.addWidget(g3)

        g4 = QGroupBox("Analisis")
        v4 = QVBoxLayout(g4)
        self._btn(v4, "Depth Profile",   self.depth_profile)
        self._btn(v4, "Ver Comparative", self.show_compare)
        self._btn(v4, "Mapping 2D",      self.map2d)
        make_collapsible(g4)
        ll.addWidget(g4)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        ll.addWidget(self.progress)
        ll.addStretch()

        right_col = QWidget()
        rc = QVBoxLayout(right_col)
        rc.setContentsMargins(0, 0, 0, 0); rc.setSpacing(4)
        rl.addWidget(right_col, stretch=1)

        self.tabs = QTabWidget(); self.tabs.setDocumentMode(True)
        rc.addWidget(self.tabs, stretch=1)

        self.tab_measure  = MeasureTab()
        self.tab_depth    = PlotTab()
        self.tab_compare  = CompareTab()
        self.tab_map      = PlotTab()

        self.tabs.addTab(self.tab_measure,  "Medida")
        self.tabs.addTab(self.tab_depth,    "Depth Profile")
        self.tabs.addTab(self.tab_compare,  "Comparative")
        self.tabs.addTab(self.tab_map,      "Mapping 2D")

        # conectar señal del boton Añadir
        self.tab_compare.add_sim_requested.connect(self._on_add_sim)
        self.tab_compare.delete_selected_requested.connect(self._on_delete_selected)
        self.tab_compare.chk_normalize.stateChanged.connect(lambda _s: self._redraw_compare())
        for w in (self.tab_compare.chk_post_sg, self.tab_compare.chk_post_clip_upper,
                  self.tab_compare.chk_post_clip_lower):
            w.stateChanged.connect(lambda _s: self._redraw_compare())
        for w in (self.tab_compare.spin_post_sg_window, self.tab_compare.spin_post_sg_poly):
            w.valueChanged.connect(lambda _v: self._redraw_compare())
        for w in (self.tab_compare.spin_post_clip_upper, self.tab_compare.spin_post_clip_lower):
            w.valueChanged.connect(lambda _v: self._redraw_compare())


        save_bar = QFrame()
        save_bar.setFrameShape(QFrame.Shape.StyledPanel)
        save_bar.setStyleSheet("QFrame{background:#1a1a2e;border-top:1px solid #444466;}")
        save_bar.setFixedHeight(46)
        sb = QHBoxLayout(save_bar)
        sb.setContentsMargins(10, 4, 10, 4); sb.setSpacing(8)
        sb.addWidget(QLabel("Nombre:"))
        self.save_name = QLineEdit("medida")
        self.save_name.setFixedWidth(140)
        self.save_name.setPlaceholderText("nombre base")
        sb.addWidget(self.save_name)
        for label, slot in [("TXT", self.save_txt),
                             ("Excel", self.save_xlsx),
                             ("ASCII", self.save_ascii)]:
            b = QPushButton(label); b.setFixedHeight(30)
            b.clicked.connect(slot); sb.addWidget(b)
        sb.addSpacing(16)
        load_btn = QPushButton("Cargar sesion")
        load_btn.setFixedHeight(30)
        load_btn.clicked.connect(self.load_session)
        sb.addWidget(load_btn)
        sb.addStretch()
        rc.addWidget(save_bar)

        self.setStatusBar(QStatusBar())
        self.log("Desconectado")

    def _btn(self, layout, text, slot, color=None):
        btn = QPushButton(text)
        if color:
            btn.setStyleSheet(f"QPushButton{{background:{color};color:white;border-radius:4px}}"
                              f"QPushButton:hover{{background:{color}bb}}")
        btn.clicked.connect(slot)
        layout.addWidget(btn)
        return btn

    def _on_filter_toggle(self, _state):
        on = self.chk_filter.isChecked()
        self.spin_filter_window.setEnabled(on)
        self.spin_filter_poly.setEnabled(on)

    def _on_clip_toggle(self, _state):
        self.spin_clip_upper.setEnabled(self.chk_clip_upper.isChecked())
        self.spin_clip_lower.setEnabled(self.chk_clip_lower.isChecked())

    def _on_filter_window_changed(self, val):
        if val % 2 == 0:
            self.spin_filter_window.blockSignals(True)
            self.spin_filter_window.setValue(val + 1)
            self.spin_filter_window.blockSignals(False)

    def _push_live_filter_params(self, *_args):
        """
        Envia los parametros de filtrado actuales (S-G + umbrales) al hilo
        de adquisicion EN CURSO, si lo hay, para que se apliquen desde el
        siguiente disparo sin necesidad de parar y reiniciar la medida.
        Si no hay ninguna medida en marcha, no hace nada (los valores se
        leeran igualmente al pulsar 'Iniciar medida' la proxima vez).
        """
        if self._trig_thread is not None and self._trig_thread.isRunning():
            self._trig_thread.update_filter_params(
                do_filter=self.chk_filter.isChecked(),
                filter_window=self.spin_filter_window.value(),
                filter_poly=self.spin_filter_poly.value(),
                clip_upper_enabled=self.chk_clip_upper.isChecked(),
                clip_upper_value=self.spin_clip_upper.value(),
                clip_lower_enabled=self.chk_clip_lower.isChecked(),
                clip_lower_value=self.spin_clip_lower.value(),
            )

    # ── gestión de espectros ─────────────────────────────────────────────
    def _append_spectrum(self, data, session):
        n     = len(self.saved) + 1
        name  = f"S{n}"
        entry = {"data": data, "name": name, "session": session}
        self.saved.append(entry)
        self._peak_cache.pop(len(self.saved) - 1, None)
        color = COLORS[(n - 1) % len(COLORS)]
        self.tab_compare.add_spectrum_row(
            name, session, color,
            on_change=self._redraw_compare,
            on_delete=self._make_delete_fn(entry)
        )

    def _make_delete_fn(self, entry_ref):
        def delete_fn(checked=False):
            for i, sp in enumerate(self.saved):
                if sp is entry_ref:
                    self._delete_spectrum(i)
                    return
        return delete_fn

    def _delete_spectrum(self, idx):
        if idx >= len(self.saved): return
        self.saved.pop(idx)
        self._peak_cache.clear()

        _, row, _, _ = self.tab_compare.checkboxes.pop(idx)
        row.deleteLater()

        self._renumber_and_relink()

        if self.saved:
            self._redraw_compare()
        else:
            self.tab_compare.ax.clear()
            self.tab_compare._style_ax()
            self.tab_compare.canvas.draw()

    def _delete_spectra_multi(self, indices):
        """Elimina varios espectros a la vez (por indices, en cualquier orden)."""
        if not indices: return
        for idx in sorted(set(indices), reverse=True):
            if idx >= len(self.saved): continue
            self.saved.pop(idx)
            _, row, _, _ = self.tab_compare.checkboxes.pop(idx)
            row.deleteLater()
        self._peak_cache.clear()
        self._renumber_and_relink()

        if self.saved:
            self._redraw_compare()
        else:
            self.tab_compare.ax.clear()
            self.tab_compare._style_ax()
            self.tab_compare.canvas.draw()

    def _renumber_and_relink(self):
        """Renombra S1..Sn, actualiza colores/etiquetas y reconecta los botones X."""
        for i, sp in enumerate(self.saved):
            sp["name"] = f"S{i + 1}"
            color = COLORS[i % len(COLORS)]
            self.tab_compare.update_row_name(i, sp["name"], color)
            cb, _, set_lbl, _ = self.tab_compare.checkboxes[i]
            cb.setToolTip(f"Conjunto: {sp['session']}")
            set_lbl.setText(sp["session"][:6])

        for i, sp in enumerate(self.saved):
            _, row, _, _ = self.tab_compare.checkboxes[i]
            del_btn = row.layout().itemAt(3).widget()
            if del_btn:
                try: del_btn.clicked.disconnect()
                except Exception: pass
                del_btn.clicked.connect(self._make_delete_fn(sp))

    def _on_delete_selected(self):
        indices = self.tab_compare.selected_for_delete_indices()
        if not indices:
            self.log("No hay espectros marcados para eliminar"); return
        n = len(indices)
        self._delete_spectra_multi(indices)
        self.log(f"Eliminados {n} espectro(s)")

    def _compute_peaks(self, idx):
        if idx not in self._peak_cache:
            s = self.saved[idx]["data"]
            thresh = 0.08 * (s.max() - s.min())
            pk, pr = find_peaks(s, prominence=thresh, distance=10)
            self._peak_cache[idx] = (pk, pr)
        return self._peak_cache[idx]

    # ── medida ───────────────────────────────────────────────────────────
    def start(self):
        if not self.specs: self.log("Conecta primero"); return
        if self._trig_thread and self._trig_thread.isRunning():
            self.log("Medida ya en curso"); return

        self.saved = []
        self._peak_cache = {}
        self.tab_compare.clear_spectra()
        self._session_count += 1

        txt = self.spin_shots.text().strip()
        n_shots    = int(txt) if txt.isdigit() else 0
        edge       = self.cmb_edge.currentIndex()
        accum_mode = self.cmb_accum.currentIndex()

        t_ms = to_ms(self.t.text(), self.cmb_t_unit.currentText())
        delay_us = to_ms(self.t_re.text(), self.cmb_delay_unit.currentText()) * 1000.0

        self._trig_thread = TriggerThread(
            specs=self.specs, t_ms=t_ms,
            n_shots=n_shots, edge=edge,
            accum_mode=accum_mode,
            integration_delay_us=delay_us,
            do_filter=self.chk_filter.isChecked(),
            filter_window=self.spin_filter_window.value(),
            filter_poly=self.spin_filter_poly.value(),
            clip_upper_enabled=self.chk_clip_upper.isChecked(),
            clip_upper_value=self.spin_clip_upper.value(),
            clip_lower_enabled=self.chk_clip_lower.isChecked(),
            clip_lower_value=self.spin_clip_lower.value(),
        )
        self._trig_thread.shot_received.connect(self._on_shot)
        self._trig_thread.finished_all.connect(self._on_finished)
        self._trig_thread.start()

        self.tab_measure.reset()
        self.tab_measure.set_state("Esperando laser...", "#50fa7b")
        self.tab_measure.set_target(n_shots)
        self.tabs.setCurrentWidget(self.tab_measure)
        shots_txt = str(n_shots) if n_shots > 0 else "inf"
        modos = ["sin acum.", "suma acum.", "promedio"]
        self.log(f"Sesion {self._session_count} - disparos: {shots_txt}  modo: {modos[accum_mode]}  "
                 f"integracion: {t_ms:.6f} ms")

    def stop(self):
        if self._trig_thread:
            self._trig_thread.stop(); self._trig_thread = None
        self.tab_measure.set_state("Parado", "#ff5555")
        self.log("Medida parada")

    def _on_shot(self, idx, wl, s):
        self.wl = wl
        self.last_spectrum = s
        base  = self.save_name.text().strip() or "medida"
        base  = "".join(c for c in base if c.isalnum() or c in "_-")
        self._append_spectrum(s.copy(), f"{base}_{self._session_count:03d}")
        self.tab_measure.update_shot(idx, wl, s, overlay=self.chk_overlay.isChecked())
        self.log(f"Disparo #{idx}")

    def _on_finished(self):
        self.tab_measure.set_state("Completado", "#f0a500")
        self.log("Secuencia completada - guardando...")
        self._trig_thread = None
        self._autosave()

    def _autosave(self):
        if not self.saved or self.wl is None: return
        import os.path
        base  = self.save_name.text().strip() or "medida"
        base  = "".join(c for c in base if c.isalnum() or c in "_-")
        fname = f"{base}_{self._session_count:03d}.txt"
        if os.path.exists(fname):
            self.log(f"AVISO: '{fname}' ya existe."); return
        try:
            with open(fname, "w") as f:
                f.write("wl\t" + "\t".join(sp["name"] for sp in self.saved) + "\n")
                for j, w in enumerate(self.wl):
                    vals = "\t".join(f"{sp['data'][j]:.6f}" for sp in self.saved)
                    f.write(f"{w:.4f}\t{vals}\n")
            self.log(f"Auto-guardado: '{fname}'  ({len(self.saved)} espectros)")
        except Exception as e:
            self.log(f"Error guardando: {e}")

    # ── simulacion LIBS desde panel de Comparative ───────────────────────
    def _on_add_sim(self):
        tc   = self.tab_compare
        item = tc.elem_list.currentItem()
        if item is None:
            self.log("Selecciona un elemento de la lista primero"); return

        element = item.text()

        # ── snapshot frozen en el momento del click ──
        Te    = tc.sim_Te.value()
        Ne    = tc.sim_Ne.value() * 1e17
        pct   = tc.sim_pct.value()
        res   = tc.sim_res.value()
        low_w = tc.sim_low.value()
        up_w  = tc.sim_up.value()
        ion   = tc.sim_ion.value()
        ws    = tc.sim_ws.currentText()
        wl_ref = self.wl.copy() if self.wl is not None else None

        self.log(f"Simulando {element} ({pct}%)  Te={Te:.2f}eV  Ne={Ne:.1e}  res={res}...")
        QApplication.processEvents()

        try:
            from SimulatedLIBS import simulation
        except ImportError:
            self.log("SimulatedLIBS no instalado — pip install SimulatedLIBS"); return

        try:
            libs = simulation.SimulatedLIBS(
                Te=Te, Ne=Ne,
                elements=[element], percentages=[100],
                resolution=res, low_w=low_w, upper_w=up_w,
                max_ion_charge=ion, webscraping=ws,
            )

            df = libs.get_interpolated_spectrum()
            wl_col  = df.columns[0]   # primera columna = longitudes de onda
            int_col = df.columns[1]   # segunda columna = intensidades
            wl_sim = df[wl_col].to_numpy(dtype=float)
            s_sim  = df[int_col].to_numpy(dtype=float)

            # se mantienen las intensidades reales simuladas por NIST/SimulatedLIBS;
            # ya NO se normaliza a [0,1] de forma forzosa.

            # interpolar al eje del espectrómetro si existe
            if wl_ref is not None:
                s_out = np.interp(wl_ref, wl_sim, s_sim)
            else:
                s_out = s_sim
                if self.wl is None:
                    self.wl = wl_sim

            # nombre = "Elemento (porcentaje%)" con parámetros congelados
            label = f"{element} ({pct}%) Te={Te:.1f} Ne={Ne:.0e}"
            self._append_spectrum(s_out.copy(), session=label)
            self.log(f"Simulado: {label}")
            self._redraw_compare()
            self.tabs.setCurrentWidget(self.tab_compare)

        except Exception as e:
            self.log(f"Error simulando {element}: {e}")

    # ── depth profile ────────────────────────────────────────────────────
    def depth_profile(self):
        if not self.saved:
            self.log("No hay espectros — haz una medida o carga una sesion"); return
        if self.wl is None or self.selected_wl is None:
            self.log("Selecciona lambda_R en Comparative primero"); return

        idx     = (np.abs(self.wl - self.selected_wl)).argmin()
        wl_real = self.wl[idx]
        visible = self.tab_compare.visible_indices()
        data    = [float(self.saved[i]["data"][idx]) for i in visible]
        labels  = [self.saved[i]["name"] for i in visible]

        ax = self.tab_depth.ax; ax.clear(); self.tab_depth._style_ax()
        ax.plot(range(1, len(data) + 1), data, marker="o", color="#f0a500",
                linewidth=1.5, markersize=5)
        ax.set_xticks(range(1, len(data) + 1))
        ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=7)
        ax.set_title(f"Depth Profile  wl = {wl_real:.2f} nm  ({len(data)} espectros)")
        ax.set_xlabel("Espectro"); ax.set_ylabel("Intensidad")
        self.tab_depth.canvas.draw()
        self.tabs.setCurrentWidget(self.tab_depth)
        self.log(f"Depth Profile en wl={wl_real:.2f} nm")

    # ── comparative ──────────────────────────────────────────────────────
    def show_compare(self):
        if not self.saved: self.log("No hay espectros guardados"); return
        self._redraw_compare()
        self.tabs.setCurrentWidget(self.tab_compare)

    def _redraw_compare(self):
        tc = self.tab_compare; ax = tc.ax
        ax.clear(); tc._style_ax()

        normalize_view = tc.chk_normalize.isChecked()
        do_post_sg    = tc.chk_post_sg.isChecked()
        post_sg_win   = tc.spin_post_sg_window.value()
        post_sg_poly  = tc.spin_post_sg_poly.value()
        clip_up_on    = tc.chk_post_clip_upper.isChecked()
        clip_up_val   = tc.spin_post_clip_upper.value()
        clip_lo_on    = tc.chk_post_clip_lower.isChecked()
        clip_lo_val   = tc.spin_post_clip_lower.value()
        visible = tc.visible_indices()
        for i in visible:
            sp    = self.saved[i]
            s_raw = sp["data"]
            if do_post_sg or clip_up_on or clip_lo_on:
                s_raw = Processor.apply_view_filters(
                    s_raw, do_sg=do_post_sg, window=post_sg_win, poly=post_sg_poly,
                    clip_upper_enabled=clip_up_on, clip_upper_value=clip_up_val,
                    clip_lower_enabled=clip_lo_on, clip_lower_value=clip_lo_val,
                )
            if normalize_view:
                smin, smax = float(s_raw.min()), float(s_raw.max())
                s = (s_raw - smin) / (smax - smin) if smax > smin else (s_raw - smin)
            else:
                s = s_raw
            color = COLORS[i % len(COLORS)]
            ax.plot(self.wl, s, label=sp["name"], color=color, linewidth=1.2)

            peaks, props = self._compute_peaks(i)
            if len(peaks) > 0:
                # se muestran TODOS los picos detectados (antes limitado a 6)
                ax.plot(self.wl[peaks], s[peaks], "v", color=color,
                        markersize=6, markeredgecolor="#ffffff44", zorder=5)
                for p in peaks:
                    ax.annotate(f"{self.wl[p]:.1f}", xy=(self.wl[p], s[p]),
                                xytext=(0, 8), textcoords="offset points",
                                ha="center", fontsize=7, color=color)

        mid = len(self.wl) // 2; third = len(self.wl) // 3
        tc._cursor_r = ax.axvline(tc._wl_r if tc._wl_r is not None else self.wl[mid],
                                  color="#ff5555", linewidth=1.2, linestyle="--", label="lambda_R")
        tc._cursor_b = ax.axvline(tc._wl_b if tc._wl_b is not None else self.wl[third],
                                  color="#8be9fd", linewidth=1.2, linestyle="--", label="lambda_B")
        ax.legend(facecolor="#1a1a2e", labelcolor="white", fontsize=8)
        ax.set_title("Comparative" + (" (normalizado 0-1)" if normalize_view else ""))
        ax.set_xlabel("Longitud de onda (nm)")
        ax.set_ylabel("Intensidad normalizada" if normalize_view else "Intensidad (cuentas)")

        for cid in tc._zoom_cids:
            tc.canvas.mpl_disconnect(cid)
        tc._zoom_cids = []

        def on_press(event):
            if event.inaxes != ax or event.xdata is None: return
            mode = tc._mode
            if mode == "view": return
            elif mode == "zoom":
                tc._zoom_start = (event.xdata, event.ydata)
                if tc._zoom_patch:
                    try: tc._zoom_patch.remove()
                    except Exception: pass
                ylo, yhi = ax.get_ylim()
                tc._zoom_patch = MplRect(
                    (event.xdata, ylo), 0, yhi - ylo,
                    linewidth=1, edgecolor="#f0a500", facecolor="#f0a50022")
                ax.add_patch(tc._zoom_patch)
            elif mode == "cursor_r":
                x = self._snap(event.xdata); tc._wl_r = x
                if tc._cursor_r: tc._cursor_r.set_xdata([x])
                self._update_selected_wl()
                self.log(f"lambda_R: {x:.2f} nm")
                tc.canvas.draw_idle(); tc.update_cursor_display()
            elif mode == "cursor_b":
                x = self._snap(event.xdata); tc._wl_b = x
                if tc._cursor_b: tc._cursor_b.set_xdata([x])
                self._update_selected_wl()
                self.log(f"lambda_B: {x:.2f} nm")
                tc.canvas.draw_idle(); tc.update_cursor_display()

        def on_motion(event):
            if event.inaxes != ax or event.xdata is None: return
            mode = tc._mode
            if mode == "zoom" and tc._zoom_start:
                x0, y0 = tc._zoom_start
                if tc._zoom_patch:
                    tc._zoom_patch.set_width(event.xdata - x0)
                    tc.canvas.draw_idle()
            elif mode in ("cursor_r", "cursor_b"):
                x = self._snap(event.xdata)
                if mode == "cursor_r" and tc._cursor_r: tc._cursor_r.set_xdata([x])
                elif mode == "cursor_b" and tc._cursor_b: tc._cursor_b.set_xdata([x])
                tc.canvas.draw_idle()

        def on_release(event):
            if tc._mode == "zoom" and tc._zoom_start and event.inaxes == ax:
                x0, y0 = tc._zoom_start
                x1 = event.xdata or x0
                xlo, xhi = min(x0, x1), max(x0, x1)
                if xhi - xlo > 1e-9:
                    ax.set_xlim(xlo, xhi)
                    self._autoscale_y_to_xrange(ax, xlo, xhi)
                if tc._zoom_patch:
                    try: tc._zoom_patch.remove()
                    except Exception: pass
                    tc._zoom_patch = None
                tc._zoom_start = None
                tc.canvas.draw_idle()

        tc._zoom_cids.append(tc.canvas.mpl_connect("button_press_event",   on_press))
        tc._zoom_cids.append(tc.canvas.mpl_connect("motion_notify_event",  on_motion))
        tc._zoom_cids.append(tc.canvas.mpl_connect("button_release_event", on_release))
        tc.canvas.draw(); tc.update_cursor_display()

    def _snap(self, x):
        return float(self.wl[(np.abs(self.wl - x)).argmin()])

    def _autoscale_y_to_xrange(self, ax, xlo, xhi, margin=0.08):
        """Ajusta el eje Y al min/max de los datos (de las lineas visibles,
        excluyendo los cursores) dentro del rango [xlo, xhi] del eje X."""
        ymin, ymax = None, None
        for line in ax.get_lines():
            label = line.get_label()
            if label in ("lambda_R", "lambda_B"):
                continue
            xdata = line.get_xdata()
            ydata = line.get_ydata()
            if len(xdata) == 0:
                continue
            mask = (xdata >= xlo) & (xdata <= xhi)
            if not np.any(mask):
                continue
            y_sel = np.asarray(ydata)[mask]
            y_sel = y_sel[np.isfinite(y_sel)]
            if y_sel.size == 0:
                continue
            lo, hi = float(y_sel.min()), float(y_sel.max())
            ymin = lo if ymin is None else min(ymin, lo)
            ymax = hi if ymax is None else max(ymax, hi)

        if ymin is None or ymax is None:
            return
        span = ymax - ymin
        if span <= 1e-12:
            span = abs(ymax) if ymax != 0 else 1.0
        pad = span * margin
        ax.set_ylim(ymin - pad, ymax + pad)

    def _update_selected_wl(self):
        tc = self.tab_compare
        if tc._wl_r is not None and tc._wl_b is not None:
            self.selected_wl = tc._wl_r - tc._wl_b
            self.log(f"lambda efectiva (Delta): {self.selected_wl:.2f} nm")
        elif tc._wl_r is not None:
            self.selected_wl = tc._wl_r

    # ── mapping 2D ───────────────────────────────────────────────────────
    def map2d(self):
        if not self.specs: self.log("Conecta primero"); return
        if self.selected_wl is None:
            self.log("Selecciona lambda_R en Comparative primero"); return
        t_ms = to_ms(self.t.text(), self.cmb_t_unit.currentText())
        self._map_thread = MapThread(
            specs=self.specs, t_ms=t_ms,
            wl=self.wl, selected_wl=self.selected_wl,
            edge=self.cmb_edge.currentIndex()
        )
        self._map_thread.progress.connect(self._on_progress)
        self._map_thread.done.connect(self._draw_map)
        self._map_thread.start()
        self.progress.setVisible(True); self.progress.setValue(0)
        self.tabs.setCurrentWidget(self.tab_map)
        self.log("Mapping 2D - esperando disparos...")

    def _draw_map(self, cube):
        self.progress.setVisible(False)
        ax = self.tab_map.ax; ax.clear(); self.tab_map._style_ax()
        im = ax.imshow(cube, origin="lower", cmap="plasma")
        self.tab_map.fig.colorbar(im, ax=ax, label="Intensidad")
        ax.set_title(f"Mapping 2D  wl = {self.selected_wl:.2f} nm")
        self.tab_map.canvas.draw()
        self.log("Mapping 2D completado")

    def _on_progress(self, val):
        self.progress.setValue(val)

    # ── carga de sesiones ────────────────────────────────────────────────
    def load_session(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "Cargar sesion(es)", "", "Text files (*.txt);;All files (*)")
        if not paths: return
        loaded = 0
        for path in paths:
            try:
                import os.path as op
                session_label = op.splitext(op.basename(path))[0]
                with open(path, "r") as f:
                    header = f.readline().strip().split("\t")
                data = np.loadtxt(path, delimiter="\t", skiprows=1)
                if data.ndim == 1 or data.shape[1] < 2:
                    self.log(f"Formato no reconocido: {path}"); continue
                wl_loaded = data[:, 0]
                spectra   = [data[:, col] for col in range(1, data.shape[1])]
                if self.wl is None:
                    self.wl = wl_loaded
                elif len(wl_loaded) != len(self.wl):
                    spectra = [np.interp(self.wl, wl_loaded, s) for s in spectra]
                for s in spectra:
                    self._append_spectrum(s.copy(), session_label)
                    loaded += 1
            except Exception as e:
                self.log(f"Error cargando {path}: {e}")
        if loaded:
            self.log(f"Cargados {loaded} espectros")
            self._redraw_compare()
            self.tabs.setCurrentWidget(self.tab_compare)

    # ── guardado ─────────────────────────────────────────────────────────
    def save_txt(self):
        if not self.saved or self.wl is None: return
        path, _ = QFileDialog.getSaveFileName(self, "Guardar TXT", "", "Text files (*.txt)")
        if not path: return
        with open(path, "w") as f:
            f.write("wl\t" + "\t".join(sp["name"] for sp in self.saved) + "\n")
            for j, w in enumerate(self.wl):
                vals = "\t".join(f"{sp['data'][j]:.6f}" for sp in self.saved)
                f.write(f"{w:.4f}\t{vals}\n")
        self.log(f"Guardado: {path}")

    def save_ascii(self):
        if not self.saved or self.wl is None: return
        path, _ = QFileDialog.getSaveFileName(self, "Guardar ASCII", "", "ASCII files (*.ascii)")
        if not path: return
        with open(path, "w") as f:
            f.write("# wl " + " ".join(sp["name"] for sp in self.saved) + "\n")
            for j, w in enumerate(self.wl):
                vals = " ".join(f"{sp['data'][j]:.6f}" for sp in self.saved)
                f.write(f"{w:.4f} {vals}\n")
        self.log(f"Guardado: {path}")

    def save_xlsx(self):
        if not self.saved or self.wl is None: return
        try:
            import openpyxl
        except ImportError:
            self.log("pip install openpyxl"); return
        path, _ = QFileDialog.getSaveFileName(self, "Guardar Excel", "", "Excel files (*.xlsx)")
        if not path: return
        if not path.endswith(".xlsx"): path += ".xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Espectros"
        ws.cell(row=1, column=1, value="wl (nm)")
        for i, sp in enumerate(self.saved):
            ws.cell(row=1, column=i + 2, value=sp["name"])
        ws.cell(row=2, column=1, value="conjunto")
        for i, sp in enumerate(self.saved):
            ws.cell(row=2, column=i + 2, value=sp["session"])
        for j, w in enumerate(self.wl):
            ws.cell(row=j + 3, column=1, value=round(float(w), 4))
            for i, sp in enumerate(self.saved):
                ws.cell(row=j + 3, column=i + 2, value=round(float(sp["data"][j]), 6))
        wb.save(path)
        self.log(f"Guardado Excel: {path}")


# ===============================================
if __name__ == "__main__":
    avaspec.AVS_Init(-1)
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = App()
    window.show()
    sys.exit(app.exec())